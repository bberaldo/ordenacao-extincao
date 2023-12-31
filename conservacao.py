#Gestão de Dados de Conservação: Desenvolver um sistema que ordena e compara dados de conservação de espécies ameaçadas, como populações de animais ou vegetação nativa, para apoiar esforços de preservação.
import openpyxl
import tkinter as tk
import time

def insertionSort(arr):
    start_time = time.time()

    for i in range(1, len(arr)):
        temp = arr[i]
        j = i - 1
        while j >= 0 and temp[0] < arr[j][0]:
            arr[j + 1] = arr[j]
            j -= 1
        arr[j + 1] = temp

    end_time = time.time()
    elapsed_time = end_time - start_time
    return arr, elapsed_time


def quicksort(arr):
    if len(arr) <= 1:
        return arr
    else:
        pivot = arr[0][0]
        left = [x for x in arr[1:] if x[0] < pivot]
        right = [x for x in arr[1:] if x[0] >= pivot]

        return quicksort(left) + [arr[0]] + quicksort(right)


def lerDados():
  try:
    wb_obj = openpyxl.load_workbook("dados.xlsx")
  except Exception as e:
    print(f'Erro: {e}')
  sheet_obj = wb_obj.active
  nRow = sheet_obj.max_row
    
  valores=[]
  for i in range(4, nRow + 1):
    cell_populacao = sheet_obj.cell(row = i, column = 5)
    cell_nome = sheet_obj.cell(row = i, column = 4)
    valores.append((cell_nome.value,int(cell_populacao.value)))
  
  return valores

def actionBtnQuickSort():
  array = lerDados()
  start_time = time.time()
  dados_ordenados = quicksort(array)
  end_time = time.time()
  elapsed_time = end_time - start_time

  # dados_ordenados = quicksort(array)
  saidaQuickSort.insert(tk.END, f"Tempo do QuickSort: {elapsed_time:.6f} segundos\n")
  for i in range(len(dados_ordenados)):
     saidaQuickSort.insert(tk.END, f"Nome: {dados_ordenados[i][0]} | População: {dados_ordenados[i][1]}")

def actionBtnInsertionSort():
  array = lerDados()
  dados_ordenados, tempo_insertion = insertionSort(array)
  saidaInsertionSort.insert(tk.END, f"Tempo do InsertionSort: {tempo_insertion:.6f} segundos\n")
  for i in range(len(dados_ordenados)):
     saidaInsertionSort.insert(tk.END, f"Nome: {dados_ordenados[i][0]} | População: {dados_ordenados[i]}")

# variáveis globais de tempo de cada algoritmo
# criando a janela
janela = tk.Tk()
janela.title("Algoritmos de Ordenação")
janela.geometry("600x400")

# Configuração do peso das linhas e colunas para centralização
janela.grid_columnconfigure(0, weight=1)
janela.grid_columnconfigure(1, weight=1)
janela.grid_rowconfigure(2, weight=1)

# Espaço vertical entre o título e os botões
espaco_vertical_titulo = tk.Label(janela, text="", pady=8)
espaco_vertical_titulo.grid(row=0, columnspan=2)

# Botão Quick Sort
btnQuickSort = tk.Button(janela, text="QuickSort", command=actionBtnQuickSort)
btnQuickSort.grid(row=1, column=0, padx=(10, 10), pady=8)

# Botão Insertion Sort
btnInsertionSort = tk.Button(janela, text="InsertionSort", command=actionBtnInsertionSort)
btnInsertionSort.grid(row=1, column=1, padx=(10, 10), pady=8)

# Espaço vertical entre os botões e as listas
espaco_vertical_entre = tk.Label(janela, text="", pady=8)
espaco_vertical_entre.grid(row=2, columnspan=2)

# Listbox para o QuickSort
saidaQuickSort = tk.Listbox(janela, height=300, width=40)
saidaQuickSort.grid(row=3, column=0, padx=(10, 10), sticky='nsew')

# Listbox para o InsertionSort
saidaInsertionSort = tk.Listbox(janela, height=300, width=40)
saidaInsertionSort.grid(row=3, column=1, padx=(10, 10), sticky='nsew')

# Iniciar loop da janela
janela.mainloop()
